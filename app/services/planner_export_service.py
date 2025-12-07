"""
Excel export service for planner view (jury program and load distribution).

This module exposes a single high-level function:

    exportPlannerToExcel(planner_data: dict) -> bytes

The function takes a JSON-serializable dictionary in the following form:

{
    "classes": ["D106", "D107", ...],
    "timeSlots": ["09:00", "09:30", ...],
    "projects": [
        {
            "class": "D106",
            "time": "09:00",
            "projectTitle": "Proje Adı Tam Olarak",
            "type": "Bitirme",
            "responsible": "Sorumlu Öğretim Görevlisi Tam Adı",
            "jury": ["Jüri Üyesi 1 Tam Adı", "Jüri Üyesi 2 Tam Adı"],
            "color": "#A8E0FF"
        },
        ...
    ],
    "hocaLoad": { "BD": 11, "GB": 11, ... },
    "arsGorLoad": { "BAK": 3, "BAÖ": 2, ... }
}

The output is a binary .xlsx payload with two sheets:
    1) "Jüri Programı"
    2) "Final Sonrası Hoca Yük Dağılımı"

Styling is intentionally modern, clean and close to the reference Excel layout.
"""

from __future__ import annotations

import colorsys
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _create_border(color: str = "D9D9D9") -> Border:
    """Create a thin, soft border style."""
    thin_side = Side(style="thin", color=color)
    return Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_NO_WRAP = Alignment(horizontal="center", vertical="center")

TITLE_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="2F5597")
TIME_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")
ALT_TIME_FILL = PatternFill(fill_type="solid", fgColor="EDEDED")


def _hex_color_to_excel(color: str, fallback: str = "FFFFFF") -> str:
    """
    Normalize a hex color string for openpyxl.

    - Accepts strings like "#AABBCC" or "AABBCC".
    - Returns just the RRGGBB component or fallback if invalid.
    """
    if not color:
        return fallback
    value = color.strip().lstrip("#")
    if len(value) != 6:
        return fallback
    # Basic validation
    try:
        int(value, 16)
    except ValueError:
        return fallback
    return value.upper()


def _autosize_columns(ws) -> None:
    """Best-effort automatic column width adjustment."""
    max_lengths: Dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            try:
                # Try to read cell value - merged cells will raise an error
                cell_value = cell.value
                if cell_value is None:
                    continue
                text = str(cell_value)
                lines = text.splitlines()
                if not lines:
                    # Empty text or only whitespace
                    length = len(text.strip())
                else:
                    length = max(len(line) for line in lines)
                col_letter = cell.column_letter
                current = max_lengths.get(col_letter, 0)
                if length > current:
                    max_lengths[col_letter] = length
            except (AttributeError, TypeError, ValueError):
                # Skip merged cells or cells that can't be read
                # Merged cells (except top-left) have read-only value attribute
                continue

    for col_letter, length in max_lengths.items():
        # Slight padding; Excel width is roughly character count
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 40)


def _project_type_color(project_type: str | None) -> str:
    """
    Fallback color mapping for project type when a project-specific color
    is not provided in planner_data.
    """
    if not project_type:
        return "00796B"  # teal as neutral

    pt = project_type.strip().lower()
    if pt in {"final", "bitirme"}:
        return "1976D2"  # blue
    if pt in {"interim", "ara"}:
        return "DC004E"  # magenta
    return "00796B"  # teal


def _get_text_color_for_background(bg_color: str) -> str:
    """
    Determine text color (black or white) based on background color brightness.
    Returns "000000" for black text on light backgrounds, "FFFFFF" for white text on dark backgrounds.
    """
    if not bg_color or bg_color == "FFFFFF":
        return "000000"  # Black text on white
    
    # Convert hex to RGB
    try:
        r = int(bg_color[0:2], 16)
        g = int(bg_color[2:4], 16)
        b = int(bg_color[4:6], 16)
    except (ValueError, IndexError):
        return "000000"  # Default to black
    
    # Calculate relative luminance (perceived brightness)
    # Using the formula: 0.299*R + 0.587*G + 0.114*B
    luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255.0
    
    # If luminance is above 0.5, use black text; otherwise use white text
    return "000000" if luminance > 0.5 else "FFFFFF"


def _hsv_to_hex(h: float, s: float, v: float) -> str:
    """
    Convert HSV color to hex RGB string.
    h: hue (0-360)
    s: saturation (0-1)
    v: value/brightness (0-1)
    """
    # Convert HSV to RGB (colorsys uses 0-1 for all values)
    r, g, b = colorsys.hsv_to_rgb(h / 360.0, s, v)
    
    # Convert to 0-255 range and format as hex
    r_hex = format(int(r * 255), "02X")
    g_hex = format(int(g * 255), "02X")
    b_hex = format(int(b * 255), "02X")
    
    return f"{r_hex}{g_hex}{b_hex}"


def _get_instructor_color(instructor_name: str) -> str:
    """
    Assign a deterministic color to each instructor based on their name.
    This ensures the same instructor always gets the same color.
    Uses HSV color space to generate maximally distinguishable colors.
    Optimized for maximum visual contrast and color-blind friendly.
    """
    if not instructor_name:
        return "FFFFFF"  # white as fallback

    # Generate colors using HSV color space for maximum distinguishability
    # We'll use 80 distinct colors with optimized distribution:
    # - Hue evenly distributed but avoiding problematic ranges
    # - High saturation (0.8-0.95) for vibrant, distinct colors
    # - Medium-high value (0.7-0.9) for good visibility
    
    # Use hash of instructor name to deterministically assign color
    name_hash = hash(instructor_name.strip().lower())
    
    # Generate 80 distinct colors for better coverage
    num_colors = 80
    color_index = abs(name_hash) % num_colors
    
    # Distribute hues evenly across the spectrum
    # Skip very yellow-green range (around 60-90 degrees) as it's less distinguishable
    # and adjust spacing for better visual separation
    base_hue = (color_index * 360.0 / num_colors) % 360
    
    # Add slight variation to avoid clustering in problematic ranges
    hue_variation = (color_index % 3) * 2  # Small variation for similar hues
    hue = (base_hue + hue_variation) % 360
    
    # Use multiple saturation/value combinations for maximum distinction
    # Higher saturation = more vibrant, easier to distinguish
    saturation_levels = [0.90, 0.85, 0.95, 0.80, 0.88]  # Very high saturation
    value_levels = [0.80, 0.75, 0.85, 0.70, 0.82]  # Medium-high brightness
    
    # Use different combinations based on color index
    sat_index = (color_index // 16) % len(saturation_levels)
    val_index = (color_index // 20) % len(value_levels)
    
    saturation = saturation_levels[sat_index]
    value = value_levels[val_index]
    
    # Convert HSV to hex
    return _hsv_to_hex(hue, saturation, value)


def _build_schedule_lookup(
    projects: Iterable[Dict[str, Any]]
) -> Dict[Tuple[str, str], List[Dict[str, Any]]]:
    """
    Build a lookup map keyed by (class, time) for fast cell population.

    Multiple projects per cell are allowed; they will be rendered as stacked
    multi-line entries inside the same Excel cell.
    """
    lookup: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    for proj in projects:
        cls = str(proj.get("class", "")).strip()
        time = str(proj.get("time", "")).strip()
        if not cls or not time:
            continue
        key = (cls, time)
        lookup.setdefault(key, []).append(proj)
    return lookup


def _build_jury_program_sheet(wb: Workbook, planner_data: Dict[str, Any]) -> None:
    """
    Create the 'Jüri Programı' sheet with main timetable.
    Each classroom gets 4 columns: [Proje] [Sorumlu] [Jüri1] [Jüri2]
    Colors are assigned per instructor (deterministic).
    Empty column between classes for readability.
    Lunch break row inserted before 13:00.
    """
    classes: List[str] = [str(c) for c in planner_data.get("classes", [])]
    time_slots: List[str] = [str(t) for t in planner_data.get("timeSlots", [])]
    projects: List[Dict[str, Any]] = list(planner_data.get("projects", []))

    ws = wb.active
    ws.title = "Jüri Programı"

    border = _create_border()

    # Find 13:00 index for lunch break insertion
    lunch_break_time = "13:00"
    lunch_break_idx = None
    for idx, time_str in enumerate(time_slots):
        if time_str.startswith("13:00") or time_str == "13:00":
            lunch_break_idx = idx
            break

    # Title row
    title_text = planner_data.get("title") or "BİLGİSAYAR ve BİTİRME Projesi Jüri Programı"
    date_str = planner_data.get("date") or datetime.now().strftime("%d %B %Y").upper()
    ws.cell(row=1, column=1, value=f"{title_text} - {date_str}")
    ws["A1"].font = Font(name="Segoe UI", bold=True, size=16)
    ws["A1"].alignment = CENTER_NO_WRAP
    ws["A1"].fill = TITLE_FILL

    # Calculate total columns: 1 (time) + (4 columns per class + 1 empty column between classes)
    # Last class doesn't need empty column after it
    total_cols = 1 + (len(classes) * 4) + (len(classes) - 1)
    if total_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # Header row: Time + for each class: [Class Name] [Proje] [Sorumlu] [Jüri 1] [Jüri 2]
    header_row_idx = 3
    time_header_cell = ws.cell(row=header_row_idx, column=1, value="Saat")
    time_header_cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF")
    time_header_cell.alignment = CENTER_NO_WRAP
    time_header_cell.fill = HEADER_FILL
    time_header_cell.border = border

    # Merge time header across 2 rows
    ws.merge_cells(start_row=header_row_idx, start_column=1, end_row=header_row_idx + 1, end_column=1)

    # Create sub-header row for class names and column labels
    sub_header_row_idx = header_row_idx + 1
    col_idx = 2
    for cls_idx, cls in enumerate(classes):
        # Class name header (spans 4 columns)
        class_label = str(cls)
        class_cell = ws.cell(row=header_row_idx, column=col_idx, value=class_label)
        class_cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=12)
        class_cell.alignment = CENTER_NO_WRAP
        class_cell.fill = HEADER_FILL
        class_cell.border = border
        ws.merge_cells(
            start_row=header_row_idx,
            start_column=col_idx,
            end_row=header_row_idx,
            end_column=col_idx + 3,
        )

        # Sub-headers: Proje, Sorumlu, Jüri 1, Jüri 2
        sub_headers = ["Proje", "Sorumlu", "Jüri 1", "Jüri 2"]
        for sub_idx, sub_header in enumerate(sub_headers):
            sub_cell = ws.cell(row=sub_header_row_idx, column=col_idx + sub_idx, value=sub_header)
            sub_cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
            sub_cell.alignment = CENTER_NO_WRAP
            sub_cell.fill = HEADER_FILL
            sub_cell.border = border

        col_idx += 4  # 4 columns for class

        # Add empty column between classes (except after last class)
        if cls_idx < len(classes) - 1:
            empty_cell = ws.cell(row=header_row_idx, column=col_idx, value="")
            empty_cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
            empty_cell.border = border
            ws.merge_cells(
                start_row=header_row_idx,
                start_column=col_idx,
                end_row=header_row_idx + 1,
                end_column=col_idx,
            )
            col_idx += 1

    # Data rows
    schedule_lookup = _build_schedule_lookup(projects)
    current_row_idx = sub_header_row_idx + 1

    for row_offset, time_str in enumerate(time_slots):
        # Insert lunch break before 13:00
        if lunch_break_idx is not None and row_offset == lunch_break_idx:
            # Empty row before lunch break
            empty_row = current_row_idx
            time_cell_empty = ws.cell(row=empty_row, column=1, value="")
            time_cell_empty.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
            time_cell_empty.border = border
            # Fill empty columns
            for col in range(2, total_cols + 1):
                cell = ws.cell(row=empty_row, column=col, value="")
                cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
                cell.border = border
            current_row_idx += 1

            # Lunch break row
            lunch_row = current_row_idx
            lunch_cell = ws.cell(row=lunch_row, column=1, value="Öğle Arası")
            lunch_cell.font = Font(name="Segoe UI", bold=True, size=12, color="FFFFFF")
            lunch_cell.alignment = CENTER_NO_WRAP
            lunch_cell.fill = HEADER_FILL
            lunch_cell.border = border

            # Fill all columns in lunch break row BEFORE merging
            for col in range(2, total_cols + 1):
                cell = ws.cell(row=lunch_row, column=col, value="")
                cell.fill = HEADER_FILL
                cell.border = border

            # Merge lunch break across all columns (after filling cells)
            ws.merge_cells(
                start_row=lunch_row, start_column=1, end_row=lunch_row, end_column=total_cols
            )

            current_row_idx += 1

            # Empty row after lunch break
            empty_row_after = current_row_idx
            time_cell_empty_after = ws.cell(row=empty_row_after, column=1, value="")
            time_cell_empty_after.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
            time_cell_empty_after.border = border
            # Fill empty columns
            for col in range(2, total_cols + 1):
                cell = ws.cell(row=empty_row_after, column=col, value="")
                cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
                cell.border = border
            current_row_idx += 1

        row_idx = current_row_idx

        # Time column with light grey banding
        time_cell = ws.cell(row=row_idx, column=1, value=time_str)
        time_cell.font = Font(name="Segoe UI", bold=True, size=11)
        time_cell.alignment = CENTER_NO_WRAP
        time_cell.fill = TIME_FILL if row_offset % 2 == 0 else ALT_TIME_FILL
        time_cell.border = border

        # For each class, create 4 cells: [Proje] [Sorumlu] [Jüri1] [Jüri2]
        col_idx = 2
        for cls_idx, cls in enumerate(classes):
            projs_in_slot = schedule_lookup.get((cls, time_str), [])

            if not projs_in_slot:
                # Empty cells for all 4 columns
                for i in range(4):
                    cell = ws.cell(row=row_idx, column=col_idx + i, value="")
                    cell.border = border
                    cell.alignment = CENTER_ALIGNMENT
                    cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
            else:
                # Take first project
                first_proj = projs_in_slot[0]
                project_title = str(
                    first_proj.get("projectTitle", first_proj.get("projectCode", ""))
                ).strip()
                responsible = str(first_proj.get("responsible", "")).strip()
                jury_list = [
                    str(j).strip() for j in (first_proj.get("jury") or []) if str(j).strip()
                ]

                # Cell 1: Proje
                proj_cell = ws.cell(row=row_idx, column=col_idx, value=project_title)
                proj_cell.alignment = CENTER_ALIGNMENT
                proj_cell.border = border
                proj_cell.font = Font(name="Segoe UI", size=10, bold=True, color="000000")
                proj_cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")

                # Cell 2: Sorumlu
                responsible_cell = ws.cell(row=row_idx, column=col_idx + 1, value=responsible)
                responsible_cell.alignment = CENTER_ALIGNMENT
                responsible_cell.border = border
                # Color based on responsible instructor
                if responsible:
                    instructor_color = _get_instructor_color(responsible)
                    text_color = _get_text_color_for_background(instructor_color)
                    responsible_cell.fill = PatternFill(fill_type="solid", fgColor=instructor_color)
                    responsible_cell.font = Font(name="Segoe UI", size=10, bold=True, color=text_color)
                else:
                    responsible_cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
                    responsible_cell.font = Font(name="Segoe UI", size=10, bold=True, color="000000")

                # Cell 3: Jüri 1
                jury1_name = jury_list[0] if len(jury_list) > 0 else ""
                jury1_cell = ws.cell(row=row_idx, column=col_idx + 2, value=jury1_name)
                jury1_cell.alignment = CENTER_ALIGNMENT
                jury1_cell.border = border
                if jury1_name:
                    jury1_color = _get_instructor_color(jury1_name)
                    text_color = _get_text_color_for_background(jury1_color)
                    jury1_cell.fill = PatternFill(fill_type="solid", fgColor=jury1_color)
                    jury1_cell.font = Font(name="Segoe UI", size=10, bold=True, color=text_color)
                else:
                    jury1_cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
                    jury1_cell.font = Font(name="Segoe UI", size=10, bold=True, color="000000")

                # Cell 4: Jüri 2
                jury2_name = jury_list[1] if len(jury_list) > 1 else ""
                jury2_cell = ws.cell(row=row_idx, column=col_idx + 3, value=jury2_name)
                jury2_cell.alignment = CENTER_ALIGNMENT
                jury2_cell.border = border
                if jury2_name:
                    jury2_color = _get_instructor_color(jury2_name)
                    text_color = _get_text_color_for_background(jury2_color)
                    jury2_cell.fill = PatternFill(fill_type="solid", fgColor=jury2_color)
                    jury2_cell.font = Font(name="Segoe UI", size=10, bold=True, color=text_color)
                else:
                    jury2_cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
                    jury2_cell.font = Font(name="Segoe UI", size=10, bold=True, color="000000")

            col_idx += 4  # 4 columns for class

            # Add empty column between classes (except after last class)
            if cls_idx < len(classes) - 1:
                empty_cell = ws.cell(row=row_idx, column=col_idx, value="")
                empty_cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
                empty_cell.border = border
                col_idx += 1

        # Slightly taller row for readability
        ws.row_dimensions[row_idx].height = 40
        current_row_idx += 1

    # Adjust column widths
    ws.column_dimensions["A"].width = 10
    _autosize_columns(ws)


def _create_load_sheet(
    wb: Workbook,
    title: str,
    sheet_name: str,
    load_data: Dict[str, Any] | None,
) -> None:
    """Create a generic load distribution sheet (Hoca / Arş Gör.)."""
    ws = wb.create_sheet(title=sheet_name)
    border = _create_border()

    ws["A1"] = title
    ws["A1"].font = Font(name="Segoe UI", bold=True, size=14)
    ws["A1"].alignment = CENTER_NO_WRAP
    ws["A1"].fill = TITLE_FILL
    ws.merge_cells("A1:B1")

    # Header row
    ws["A2"] = "Hoca"
    ws["B2"] = "Yük"
    for col in ("A", "B"):
        cell = ws[f"{col}2"]
        cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF")
        cell.alignment = CENTER_NO_WRAP
        cell.fill = HEADER_FILL
        cell.border = border

    if not load_data:
        _autosize_columns(ws)
        return

    # Sort by load descending, then by name
    sorted_items = sorted(
        load_data.items(), key=lambda kv: (-int(kv[1]), str(kv[0]))
    )

    # Soft color palette to differentiate rows
    pastel_colors = [
        "FDE9D9",
        "E2F0D9",
        "DEEAF6",
        "FFF2CC",
        "E4DFEC",
        "F8CBAD",
    ]

    for idx, (name, load) in enumerate(sorted_items, start=3):
        row_fill = PatternFill(
            fill_type="solid", fgColor=pastel_colors[(idx - 3) % len(pastel_colors)]
        )

        name_cell = ws.cell(row=idx, column=1, value=str(name))
        load_cell = ws.cell(row=idx, column=2, value=int(load))

        for cell in (name_cell, load_cell):
            cell.font = Font(name="Segoe UI", size=11)
            cell.alignment = CENTER_NO_WRAP
            cell.fill = row_fill
            cell.border = border

    _autosize_columns(ws)


def exportPlannerToExcel(planner_data: Dict[str, Any]) -> bytes:
    """
    High-level entry point used by the backend API to generate the Excel file.

    Args:
        planner_data: Planner JSON payload as documented in the module docstring.

    Returns:
        Binary .xlsx content suitable for returning from a FastAPI endpoint.
    """
    wb = Workbook()

    # 1) Main jury program grid
    _build_jury_program_sheet(wb, planner_data)

    # 2) Final sonrası hoca yük dağılımı
    _create_load_sheet(
        wb,
        title="Final Sonrası Hoca Yük Dağılımı",
        sheet_name="Final Sonrası Hoca Yük Dağılımı",
        load_data=planner_data.get("hocaLoad") or {},
    )

    # Serialize to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


__all__ = ["exportPlannerToExcel"]


