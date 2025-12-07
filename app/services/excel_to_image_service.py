"""
Service for converting Excel files to PNG images.
Used for embedding planner calendars in email notifications.
"""

import logging
from io import BytesIO
from typing import Optional
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from PIL import Image, ImageDraw, ImageFont
import math

logger = logging.getLogger(__name__)


def excel_to_png(excel_bytes: bytes, sheet_name: Optional[str] = None, max_width: int = 2400) -> bytes:
    """
    Convert Excel workbook to PNG image.
    
    Args:
        excel_bytes: Binary Excel file content (.xlsx)
        sheet_name: Name of sheet to convert (defaults to first sheet)
        max_width: Maximum width of output image in pixels
        
    Returns:
        PNG image bytes
    """
    try:
        # Load workbook from bytes
        logger.info(f"Loading Excel workbook from {len(excel_bytes)} bytes")
        # Use data_only=False to get actual cell values (not calculated values)
        workbook = load_workbook(BytesIO(excel_bytes), data_only=False)
        logger.info(f"Workbook loaded. Available sheets: {workbook.sheetnames}")
        
        # Select sheet
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found in {workbook.sheetnames}, using first sheet: {workbook.active.title}")
                sheet = workbook.active
            else:
                sheet = workbook[sheet_name]
                logger.info(f"Using sheet: {sheet_name}")
        else:
            sheet = workbook.active
            logger.info(f"Using default sheet: {sheet.title}")
        
        # Get sheet dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column
        logger.info(f"Sheet dimensions: {max_row} rows x {max_col} columns")
        
        if max_row == 0 or max_col == 0:
            logger.warning("Empty sheet, creating placeholder image")
            img = Image.new('RGB', (800, 200), color='white')
            draw = ImageDraw.Draw(img)
            try:
                font = ImageFont.truetype("arial.ttf", 24)
            except:
                font = ImageFont.load_default()
            draw.text((20, 80), "No data available", fill='black', font=font)
            buffer = BytesIO()
            img.save(buffer, format='PNG')
            buffer.seek(0)
            return buffer.getvalue()
        
        # Calculate cell dimensions
        # Estimate cell width and height based on content
        cell_width = 120  # Base width
        cell_height = 40  # Base height
        
        # Adjust for content
        for row in sheet.iter_rows(min_row=1, max_row=min(max_row, 100), min_col=1, max_col=max_col):
            for cell in row:
                if cell.value:
                    text = str(cell.value)
                    # Estimate width based on text length
                    estimated_width = len(text) * 7 + 20
                    if estimated_width > cell_width:
                        cell_width = min(estimated_width, 300)
        
        # Calculate image dimensions
        total_width = max_col * cell_width
        total_height = max_row * cell_height
        
        # Scale down if too wide
        scale = 1.0
        if total_width > max_width:
            scale = max_width / total_width
            cell_width = int(cell_width * scale)
            cell_height = int(cell_height * scale)
            total_width = max_col * cell_width
            total_height = max_row * cell_height
        
        # Create image
        img = Image.new('RGB', (total_width, total_height), color='white')
        draw = ImageDraw.Draw(img)
        
        # Try to load a nice font
        try:
            font_normal = ImageFont.truetype("arial.ttf", int(11 * scale))
            font_bold = ImageFont.truetype("arialbd.ttf", int(11 * scale))
            font_small = ImageFont.truetype("arial.ttf", int(9 * scale))
        except:
            try:
                font_normal = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", int(11 * scale))
                font_bold = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", int(11 * scale))
                font_small = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", int(9 * scale))
            except:
                font_normal = ImageFont.load_default()
                font_bold = ImageFont.load_default()
                font_small = ImageFont.load_default()
        
        # Draw cells
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col), start=1):
            for col_idx, cell in enumerate(row, start=1):
                x = (col_idx - 1) * cell_width
                y = (row_idx - 1) * cell_height
                
                # Get cell style
                fill_color = (255, 255, 255)  # white RGB tuple
                text_color = (0, 0, 0)  # black RGB tuple
                is_bold = False
                
                def hex_to_rgb(hex_str):
                    """Convert hex color string to RGB tuple."""
                    if not hex_str:
                        return (255, 255, 255)
                    # Convert to string if it's an RGB object
                    hex_str = str(hex_str).lstrip('#')
                    # Remove 'FF' prefix if present (ARGB format)
                    if hex_str.startswith('FF') and len(hex_str) == 8:
                        hex_str = hex_str[2:]
                    if len(hex_str) == 6:
                        try:
                            return tuple(int(hex_str[i:i+2], 16) for i in (0, 2, 4))
                        except ValueError:
                            return (255, 255, 255)
                    return (255, 255, 255)  # default white
                
                def get_rgb_from_color(color_obj):
                    """Extract RGB tuple from openpyxl color object."""
                    if not color_obj:
                        return None
                    try:
                        rgb_value = color_obj.rgb
                        if rgb_value is None:
                            return None
                        # Handle RGB object (convert to string)
                        if hasattr(rgb_value, '__str__'):
                            rgb_str = str(rgb_value)
                        else:
                            rgb_str = rgb_value
                        return hex_to_rgb(rgb_str)
                    except Exception as e:
                        logger.debug(f"Error extracting RGB from color: {e}")
                        return None
                
                if cell.fill and isinstance(cell.fill, PatternFill):
                    if cell.fill.start_color:
                        extracted_color = get_rgb_from_color(cell.fill.start_color)
                        if extracted_color:
                            fill_color = extracted_color
                
                if cell.font:
                    if cell.font.color:
                        extracted_color = get_rgb_from_color(cell.font.color)
                        if extracted_color:
                            text_color = extracted_color
                    is_bold = cell.font.bold or False
                
                # Draw cell background
                outline_color = (217, 217, 217)  # #D9D9D9 RGB
                draw.rectangle(
                    [x, y, x + cell_width - 1, y + cell_height - 1],
                    fill=fill_color,
                    outline=outline_color
                )
                
                # Draw cell text
                # Check merged cells - if cell is part of a merged range, get the value from the top-left cell
                cell_value = cell.value
                if cell_value is None:
                    # Check if this cell is part of a merged range
                    for merged_range in sheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            # Get the top-left cell of the merged range
                            top_left_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
                            cell_value = top_left_cell.value
                            break
                
                if cell_value:
                    text = str(cell_value)
                    # Handle multi-line text
                    lines = text.split('\n')
                    font_to_use = font_bold if is_bold else font_normal
                    
                    # Calculate text position (centered)
                    line_height = int(12 * scale)
                    total_text_height = len(lines) * line_height
                    start_y = y + (cell_height - total_text_height) // 2
                    
                    for line_idx, line in enumerate(lines[:5]):  # Max 5 lines per cell (increased from 3)
                        if not line.strip():
                            continue
                        text_y = start_y + line_idx * line_height
                        
                        # Truncate if too long
                        max_chars = int(cell_width / (6 * scale))  # Slightly more characters per line
                        if len(line) > max_chars:
                            line = line[:max_chars-3] + "..."
                        
                        # Draw text
                        try:
                            draw.text(
                                (x + cell_width // 2, text_y),
                                line,
                                fill=text_color,
                                font=font_to_use,
                                anchor='mm'  # Middle-center anchor
                            )
                        except Exception as e:
                            logger.debug(f"Error drawing text in cell {cell.coordinate}: {e}")
                            # Fallback: draw without anchor
                            try:
                                bbox = draw.textbbox((0, 0), line, font=font_to_use)
                                text_width = bbox[2] - bbox[0]
                                text_x = x + (cell_width - text_width) // 2
                                draw.text((text_x, text_y), line, fill=text_color, font=font_to_use)
                            except:
                                pass  # Skip if still fails
        
        # Save to bytes
        buffer = BytesIO()
        try:
            img.save(buffer, format='PNG', optimize=False)  # optimize=False for better compatibility
            buffer.seek(0)
            png_data = buffer.getvalue()
            if not png_data or len(png_data) < 100:  # PNG files should be at least 100 bytes
                raise ValueError(f"Generated PNG is too small: {len(png_data)} bytes")
            logger.info(f"Successfully converted Excel to PNG: {len(png_data)} bytes, dimensions: {total_width}x{total_height}")
            return png_data
        except Exception as save_error:
            logger.error(f"Error saving PNG image: {str(save_error)}")
            raise
        
    except Exception as e:
        logger.error(f"Error converting Excel to PNG: {str(e)}", exc_info=True)
        # Return placeholder image on error
        try:
            img = Image.new('RGB', (800, 200), color='white')
            draw = ImageDraw.Draw(img)
            try:
                font = ImageFont.truetype("arial.ttf", 24)
            except:
                font = ImageFont.load_default()
            draw.text((20, 80), f"Error generating image: {str(e)}", fill='red', font=font)
            buffer = BytesIO()
            img.save(buffer, format='PNG')
            return buffer.getvalue()
        except Exception as e2:
            logger.error(f"Error creating placeholder image: {str(e2)}")
            # Return minimal valid PNG (1x1 transparent pixel)
            return b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4\x89\x00\x00\x00\nIDATx\x9cc\x00\x01\x00\x00\x05\x00\x01\r\n-\xdb\x00\x00\x00\x00IEND\xaeB`\x82'

