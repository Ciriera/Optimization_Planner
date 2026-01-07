"""
Excel Import Service for importing instructor-project assignments from Excel files.
"""

import logging
from typing import Dict, List, Any, Optional, Tuple
from io import BytesIO
from datetime import datetime
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload

from app.models.project import Project, ProjectType, ProjectStatus
from app.models.instructor import Instructor, InstructorType
from app.db.base_class import Base

logger = logging.getLogger(__name__)


class ImportService:
    """Service for importing instructor-project data from Excel files."""

    def __init__(self):
        self.required_columns = ["InstructorName", "ProjectType", "ProjectDescription"]
        self.email_required_columns = ["InstructorName", "Email"]
        self.project_type_mapping = {
            "ara proje": ProjectType.INTERIM,
            "ara projesi": ProjectType.INTERIM,
            "interim": ProjectType.INTERIM,
            "bitirme projesi": ProjectType.FINAL,
            "bitirme": ProjectType.FINAL,
            "final": ProjectType.FINAL,
        }

    # ------------------------------------------------------------------
    # Instructor Email Template & Import
    # ------------------------------------------------------------------

    async def generate_instructor_email_template(self, db: AsyncSession) -> bytes:
        """
        Export current instructors with their emails (if any) as Excel template.
        Columns: InstructorName, Email
        """
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Instructors"
        ws.append(self.email_required_columns)

        # Fetch instructors
        result = await db.execute(select(Instructor).order_by(Instructor.name))
        instructors = result.scalars().all()

        for inst in instructors:
            ws.append([inst.name, inst.email or ""])

        stream = BytesIO()
        wb.save(stream)
        return stream.getvalue()

    def _validate_email_format(self, email: str) -> bool:
        if not email:
            return True  # allow empty (means no update)
        email_regex = r"^[\\w\\.-]+@[\\w\\.-]+\\.[a-zA-Z]{2,}$"
        return re.match(email_regex, email) is not None

    async def validate_instructor_emails(self, file_content: bytes, db: AsyncSession) -> Dict[str, Any]:
        """
        Validate instructor email Excel file.
        """
        try:
            wb = load_workbook(filename=BytesIO(file_content), data_only=True)
            sheet = wb.active

            header_row = self._find_header_row(sheet, required=self.email_required_columns)
            if header_row is None:
                return {
                    "success": False,
                    "error": f"Could not find header row. Expected columns: {', '.join(self.email_required_columns)}",
                    "rows": [],
                    "total_rows": 0,
                    "valid_rows": 0,
                    "invalid_rows": 0,
                }

            headers = [cell.value for cell in sheet[header_row]]
            header_validation = self._validate_headers(headers, required=self.email_required_columns)
            if not header_validation["valid"]:
                return {
                    "success": False,
                    "error": header_validation["error"],
                    "rows": [],
                    "total_rows": 0,
                    "valid_rows": 0,
                    "invalid_rows": 0,
                }

            column_map = self._map_columns(headers, required=self.email_required_columns)

            result = await db.execute(select(Instructor))
            instructors = result.scalars().all()
            instructor_lookup = {inst.name.lower(): inst for inst in instructors}

            rows = []
            total_rows = 0
            valid_rows = 0
            invalid_rows = 0

            for row_idx in range(header_row + 1, sheet.max_row + 1):
                instructor_name = sheet.cell(row=row_idx, column=column_map["InstructorName"]).value
                email_val = sheet.cell(row=row_idx, column=column_map["Email"]).value

                if not instructor_name and not email_val:
                    continue

                total_rows += 1
                row_errors = []

                name_norm = self._normalize_text(instructor_name) if instructor_name else None
                email_norm = email_val.strip() if isinstance(email_val, str) else email_val

                if not name_norm:
                    row_errors.append("InstructorName is required")
                inst = instructor_lookup.get(name_norm or "")
                if not inst:
                    row_errors.append(f"Instructor not found: {instructor_name}")

                if email_norm and not self._validate_email_format(email_norm):
                    row_errors.append(f"Invalid email: {email_norm}")

                row_data = {
                    "row_number": row_idx,
                    "instructor_name": instructor_name,
                    "email": email_norm or "",
                    "instructor_id": inst.id if inst else None,
                    "errors": row_errors,
                }

                if row_errors:
                    invalid_rows += 1
                else:
                    valid_rows += 1

                rows.append(row_data)

            success = invalid_rows == 0
            return {
                "success": success,
                "rows": rows,
                "total_rows": total_rows,
                "valid_rows": valid_rows,
                "invalid_rows": invalid_rows,
                "error": None if success else "Some rows contain errors",
            }
        except Exception as e:
            logger.error(f"validate_instructor_emails error: {e}", exc_info=True)
            return {
                "success": False,
                "error": str(e),
                "rows": [],
                "total_rows": 0,
                "valid_rows": 0,
                "invalid_rows": 0,
            }

    async def execute_instructor_email_import(self, file_content: bytes, db: AsyncSession, dry_run: bool = False) -> Dict[str, Any]:
        """
        Import instructor emails (update Instructor.email).
        """
        validation = await self.validate_instructor_emails(file_content, db)
        if not validation.get("success"):
            return {
                "success": False,
                "error": validation.get("error"),
                "statistics": {},
                "errors": [row for row in validation.get("rows", []) if row.get("errors")],
                "dry_run": dry_run,
                "timestamp": datetime.now().isoformat(),
            }

        updated = 0
        errors = []

        if not dry_run:
            for row in validation.get("rows", []):
                if row.get("errors"):
                    errors.append(row)
                    continue
                inst_id = row.get("instructor_id")
                if not inst_id:
                    continue
                email_val = row.get("email") or None
                await db.execute(
                    Instructor.__table__.update()
                    .where(Instructor.id == inst_id)
                    .values(email=email_val)
                )
                updated += 1
            await db.commit()

        return {
            "success": True,
            "error": None,
            "statistics": {
                "updated": updated,
                "total_rows": validation.get("total_rows", 0),
                "valid_rows": validation.get("valid_rows", 0),
                "invalid_rows": validation.get("invalid_rows", 0),
            },
            "errors": errors,
            "dry_run": dry_run,
            "timestamp": datetime.now().isoformat(),
        }

    def parse_excel_file(self, file_content: bytes) -> Dict[str, Any]:
        """
        Parse Excel file and extract instructor-project data.
        
        Args:
            file_content: Excel file content as bytes
            
        Returns:
            Dictionary with parsed data and validation results
        """
        try:
            workbook = load_workbook(filename=BytesIO(file_content), data_only=True)
            sheet = workbook.active
            
            # Find header row
            header_row = self._find_header_row(sheet)
            if header_row is None:
                return {
                    "success": False,
                    "error": "Could not find header row. Expected columns: InstructorName, ProjectType, ProjectDescription",
                    "rows": [],
                }
            
            # Validate headers
            headers = [cell.value for cell in sheet[header_row]]
            header_validation = self._validate_headers(headers)
            if not header_validation["valid"]:
                return {
                    "success": False,
                    "error": header_validation["error"],
                    "rows": [],
                }
            
            # Map column indices
            column_map = self._map_columns(headers)
            
            # Parse rows
            parsed_rows = []
            errors = []
            
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                row_data = {}
                row_errors = []
                
                # Get cell values
                instructor_name = sheet.cell(row=row_idx, column=column_map["InstructorName"]).value
                project_type = sheet.cell(row=row_idx, column=column_map["ProjectType"]).value
                project_description = sheet.cell(row=row_idx, column=column_map["ProjectDescription"]).value
                
                # Skip empty rows
                if not instructor_name and not project_type and not project_description:
                    continue
                
                # Validate and normalize
                instructor_name = self._normalize_text(instructor_name) if instructor_name else None
                project_type = self._normalize_text(project_type) if project_type else None
                project_description = self._normalize_text(project_description) if project_description else None
                
                # Validate required fields
                if not instructor_name:
                    row_errors.append("InstructorName is required")
                if not project_type:
                    row_errors.append("ProjectType is required")
                if not project_description:
                    row_errors.append("ProjectDescription is required")
                
                # Validate project type
                normalized_type = self._normalize_project_type(project_type) if project_type else None
                if project_type and not normalized_type:
                    row_errors.append(f"Invalid ProjectType: '{project_type}'. Expected: Ara Proje, Bitirme Projesi")
                
                row_data = {
                    "row_number": row_idx,
                    "instructor_name": instructor_name,
                    "project_type": project_type,
                    "project_type_normalized": normalized_type,
                    "project_description": project_description,
                    "errors": row_errors,
                    "warnings": [],
                }
                
                if row_errors:
                    errors.extend([f"Row {row_idx}: {err}" for err in row_errors])
                else:
                    # Generate project title
                    row_data["project_title"] = self._generate_project_title(
                        project_description, normalized_type
                    )
                
                parsed_rows.append(row_data)
            
            return {
                "success": len(errors) == 0,
                "error": "; ".join(errors) if errors else None,
                "rows": parsed_rows,
                "total_rows": len(parsed_rows),
                "valid_rows": len([r for r in parsed_rows if not r.get("errors")]),
                "invalid_rows": len([r for r in parsed_rows if r.get("errors")]),
            }
            
        except Exception as e:
            logger.error(f"Error parsing Excel file: {str(e)}", exc_info=True)
            return {
                "success": False,
                "error": f"Error parsing Excel file: {str(e)}",
                "rows": [],
            }

    async def validate_import_data(
        self, parsed_data: Dict[str, Any], db: AsyncSession
    ) -> Dict[str, Any]:
        """
        Validate parsed data against database.
        
        Args:
            parsed_data: Parsed data from parse_excel_file
            db: Database session
            
        Returns:
            Validation results with status for each row
        """
        if not parsed_data.get("success"):
            return parsed_data
        
        validated_rows = []
        instructors_cache = {}
        
        # Get all existing instructors
        result = await db.execute(select(Instructor))
        existing_instructors = result.scalars().all()
        for inst in existing_instructors:
            instructors_cache[self._normalize_text(inst.name)] = inst
        
        # Get all existing projects
        result = await db.execute(select(Project))
        existing_projects = result.scalars().all()
        projects_cache = {}
        for proj in existing_projects:
            key = (
                self._normalize_text(proj.title),
                proj.type,
                self._normalize_text(proj.description or ""),
            )
            projects_cache[key] = proj
        
        for row in parsed_data["rows"]:
            if row.get("errors"):
                validated_rows.append({
                    **row,
                    "status": "error",
                    "instructor_status": None,
                    "project_status": None,
                })
                continue
            
            # Check instructor
            instructor_name = row["instructor_name"]
            instructor_key = self._normalize_text(instructor_name)
            instructor = instructors_cache.get(instructor_key)
            
            if instructor:
                instructor_status = "existing"
                instructor_id = instructor.id
            else:
                instructor_status = "new"
                instructor_id = None
            
            # Check project
            project_key = (
                self._normalize_text(row["project_title"]),
                row["project_type_normalized"],
                self._normalize_text(row["project_description"]),
            )
            project = projects_cache.get(project_key)
            
            if project:
                project_status = "existing"
                project_id = project.id
            else:
                project_status = "new"
                project_id = None
            
            # Determine overall status
            if row.get("errors"):
                status = "error"
            elif instructor_status == "new" and project_status == "new":
                status = "new"
            elif instructor_status == "existing" and project_status == "new":
                status = "new_project"
            elif project_status == "existing":
                status = "duplicate"
            else:
                status = "ready"
            
            validated_rows.append({
                **row,
                "status": status,
                "instructor_status": instructor_status,
                "project_status": project_status,
                "instructor_id": instructor_id,
                "project_id": project_id,
            })
        
        return {
            **parsed_data,
            "rows": validated_rows,
        }

    async def execute_import(
        self, validated_data: Dict[str, Any], db: AsyncSession, dry_run: bool = False
    ) -> Dict[str, Any]:
        """
        Execute import: create instructors and projects in database.
        
        Args:
            validated_data: Validated data from validate_import_data
            db: Database session
            dry_run: If True, don't actually save to database
            
        Returns:
            Import results with statistics
        """
        if not validated_data.get("success"):
            return {
                "success": False,
                "error": validated_data.get("error", "Validation failed"),
                "statistics": {},
            }
        
        statistics = {
            "instructors_created": 0,
            "instructors_existing": 0,
            "projects_created": 0,
            "projects_existing": 0,
            "rows_processed": 0,
            "rows_skipped": 0,
            "rows_errors": 0,
        }
        
        instructors_cache = {}
        projects_created = []
        errors = []
        
        try:
            # Get all existing instructors
            result = await db.execute(select(Instructor))
            existing_instructors = result.scalars().all()
            for inst in existing_instructors:
                key = self._normalize_text(inst.name)
                instructors_cache[key] = inst
            
            for row in validated_data["rows"]:
                if row.get("errors") or row.get("status") == "error":
                    statistics["rows_errors"] += 1
                    errors.append(f"Row {row['row_number']}: {', '.join(row.get('errors', []))}")
                    continue
                
                if row.get("status") == "duplicate":
                    statistics["rows_skipped"] += 1
                    continue
                
                statistics["rows_processed"] += 1
                
                # Handle instructor
                instructor_name = row["instructor_name"]
                instructor_key = self._normalize_text(instructor_name)
                
                if instructor_key not in instructors_cache:
                    # Create new instructor
                    if not dry_run:
                        new_instructor = Instructor(
                            name=instructor_name,
                            type=InstructorType.INSTRUCTOR.value,
                            email=None,
                            bitirme_count=0,
                            ara_count=0,
                            total_load=0,
                        )
                        db.add(new_instructor)
                        await db.flush()
                        instructors_cache[instructor_key] = new_instructor
                        statistics["instructors_created"] += 1
                    else:
                        # Create a temporary object for dry run
                        class TempInstructor:
                            def __init__(self, name):
                                self.id = None
                                self.name = name
                        instructors_cache[instructor_key] = TempInstructor(instructor_name)
                        statistics["instructors_created"] += 1
                else:
                    statistics["instructors_existing"] += 1
                
                instructor = instructors_cache[instructor_key]
                
                # Handle project
                project_title = row["project_title"]
                project_description = row["project_description"]
                project_type = row["project_type_normalized"]
                
                # Check if project already exists
                project_key = (
                    self._normalize_text(project_title),
                    project_type,
                    self._normalize_text(project_description),
                )
                
                # Check in database
                result = await db.execute(
                    select(Project).where(
                        func.lower(Project.title) == project_title.lower(),
                        Project.type == project_type,
                    )
                )
                existing_project = result.scalar_one_or_none()
                
                # Also check in projects_created list (for same import session)
                if not existing_project:
                    for created_id in projects_created:
                        created_proj = await db.get(Project, created_id)
                        if created_proj and created_proj.title.lower() == project_title.lower() and created_proj.type == project_type:
                            existing_project = created_proj
                            break
                
                if existing_project:
                    statistics["projects_existing"] += 1
                    # Update responsible instructor if different
                    if existing_project.responsible_instructor_id != instructor.id:
                        if not dry_run:
                            existing_project.responsible_instructor_id = instructor.id
                            await db.flush()
                else:
                    # Create new project
                    if not dry_run:
                        new_project = Project(
                            title=project_title,
                            description=project_description,
                            type=project_type,
                            status=ProjectStatus.ACTIVE,
                            responsible_instructor_id=instructor.id,
                            student_capacity=1,
                            is_active=True,
                            is_makeup=False,
                        )
                        db.add(new_project)
                        await db.flush()
                        projects_created.append(new_project.id)
                        statistics["projects_created"] += 1
                    else:
                        statistics["projects_created"] += 1
            
            if not dry_run:
                await db.commit()
                logger.info(f"Import completed: {statistics}")
            else:
                await db.rollback()
                logger.info(f"Dry run completed: {statistics}")
            
            return {
                "success": True,
                "statistics": statistics,
                "errors": errors if errors else None,
                "projects_created_ids": projects_created if not dry_run else [],
            }
            
        except Exception as e:
            await db.rollback()
            logger.error(f"Error executing import: {str(e)}", exc_info=True)
            return {
                "success": False,
                "error": f"Error executing import: {str(e)}",
                "statistics": statistics,
            }

    def _find_header_row(self, sheet, required: Optional[List[str]] = None) -> Optional[int]:
        """Find the row containing headers."""
        required_cols = required or self.required_columns
        for row_idx in range(1, min(10, sheet.max_row + 1)):  # Check first 10 rows
            row_values = [cell.value for cell in sheet[row_idx]]
            row_lower = [str(v).lower() if v else "" for v in row_values]
            
            # Check if this row contains all required headers
            found_headers = []
            for header in required_cols:
                header_lower = header.lower()
                for val in row_lower:
                    if header_lower in val or val in header_lower:
                        found_headers.append(header)
                        break
            
            if len(found_headers) >= len(required_cols):
                return row_idx
        
        return None

    def _validate_headers(self, headers: List[Any], required: Optional[List[str]] = None) -> Dict[str, Any]:
        """Validate that all required headers are present."""
        required_cols = required or self.required_columns
        headers_lower = [str(h).lower() if h else "" for h in headers]
        missing = []
        
        for required in required_cols:
            required_lower = required.lower()
            found = False
            for header in headers_lower:
                if required_lower in header or header in required_lower:
                    found = True
                    break
            if not found:
                missing.append(required)
        
        if missing:
            return {
                "valid": False,
                "error": f"Missing required columns: {', '.join(missing)}",
            }
        
        return {"valid": True}

    def _map_columns(self, headers: List[Any], required: Optional[List[str]] = None) -> Dict[str, int]:
        """Map column names to column indices (1-based)."""
        required_cols = required or self.required_columns
        column_map = {}
        headers_lower = [str(h).lower() if h else "" for h in headers]
        
        for required in required_cols:
            required_lower = required.lower()
            for idx, header in enumerate(headers_lower, start=1):
                if required_lower in header or header in required_lower:
                    column_map[required] = idx
                    break
        
        return column_map

    def _normalize_text(self, text: Any) -> Optional[str]:
        """Normalize text: trim, lowercase."""
        if text is None:
            return None
        if isinstance(text, (int, float)):
            text = str(text)
        return str(text).strip()

    def _normalize_project_type(self, project_type: str) -> Optional[ProjectType]:
        """Normalize project type string to ProjectType enum."""
        if not project_type:
            return None
        
        normalized = project_type.lower().strip()
        return self.project_type_mapping.get(normalized)

    def _generate_project_title(self, description: str, project_type: ProjectType) -> str:
        """Generate project title from description."""
        if not description:
            return ""
        
        # Use description as title, or generate from type
        title = description.strip()
        
        # If title is too short, enhance it
        if len(title) < 5:
            type_name = "Ara Proje" if project_type == ProjectType.INTERIM else "Bitirme Projesi"
            title = f"{type_name} - {title}"
        
        return title

    def generate_template(self) -> bytes:
        """Generate Excel template file."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Import Template"
        
        # Headers
        headers = ["InstructorName", "ProjectType", "ProjectDescription"]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Example rows
        examples = [
            ["Eren Cavuldak", "Ara Proje", "Ara Proje 1"],
            ["Kerim Öztürk", "Bitirme Projesi", "Bitirme Projesi 5"],
            ["Ayşe Yılmaz", "Ara Proje", "Ara Proje 2"],
        ]
        
        for row_idx, example in enumerate(examples, start=2):
            for col_idx, value in enumerate(example, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 25
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

