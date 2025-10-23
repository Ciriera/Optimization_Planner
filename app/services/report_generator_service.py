"""
Report Generator Service for PDF and Excel Outputs
Proje açıklamasına göre: PDF / Excel çıktıları
"""

import os
import json
from typing import Dict, List, Any, Optional
from datetime import datetime
from pathlib import Path
import logging

# PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    print("Warning: PDF generation libraries not available. Install reportlab for PDF support.")

# Excel generation
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: Excel generation libraries not available. Install openpyxl and pandas for Excel support.")

from app.services.score_generator_service import ScoreGeneratorService
from app.services.final_makeup_service import FinalMakeupService

logger = logging.getLogger(__name__)

class ReportGeneratorService:
    """
    PDF ve Excel rapor üretici servis.
    Proje açıklamasına göre: Hangi hocanın hangi sınıf ve saat diliminde görevli olduğu + Atanan projeler + Yük dağılımı grafiği
    """
    
    def __init__(self):
        self.score_generator = ScoreGeneratorService()
        self.final_makeup_service = FinalMakeupService()
        self.output_dir = Path("app/static/reports")
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    async def generate_comprehensive_report(self, algorithm_run_id: Optional[int] = None, 
                                         format_types: List[str] = ["pdf", "excel"]) -> Dict[str, Any]:
        """
        Kapsamlı rapor oluşturur (PDF ve Excel).
        
        Args:
            algorithm_run_id: Algoritma çalıştırma ID'si
            format_types: Üretilecek format türleri ["pdf", "excel"]
            
        Returns:
            Report generation result
        """
        try:
            logger.info(f"Generating comprehensive report for algorithm run: {algorithm_run_id}")
            
            # 1. Score verilerini al
            score_result = await self.score_generator.generate_comprehensive_score_report(algorithm_run_id)
            if not score_result["success"]:
                return {
                    "success": False,
                    "error": "Score data generation failed",
                    "message": score_result.get("message", "Unknown error")
                }
            
            score_data = score_result["score_data"]
            
            # 2. Rapor verilerini hazırla
            report_data = await self._prepare_report_data(score_data, algorithm_run_id)
            
            # 3. Format türlerine göre rapor oluştur
            generated_files = []
            
            if "pdf" in format_types:
                if PDF_AVAILABLE:
                    pdf_result = await self._generate_pdf_report(report_data, algorithm_run_id)
                    if pdf_result["success"]:
                        generated_files.append({
                            "format": "pdf",
                            "file_path": pdf_result["file_path"],
                            "file_size": pdf_result["file_size"]
                        })
                else:
                    logger.warning("PDF generation skipped - reportlab not available")
            
            if "excel" in format_types:
                if EXCEL_AVAILABLE:
                    excel_result = await self._generate_excel_report(report_data, algorithm_run_id)
                    if excel_result["success"]:
                        generated_files.append({
                            "format": "excel",
                            "file_path": excel_result["file_path"],
                            "file_size": excel_result["file_size"]
                        })
                else:
                    logger.warning("Excel generation skipped - openpyxl not available")
            
            return {
                "success": True,
                "generated_files": generated_files,
                "report_data": report_data,
                "message": f"Rapor başarıyla oluşturuldu: {len(generated_files)} dosya"
            }
            
        except Exception as e:
            logger.error(f"Error generating comprehensive report: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "message": "Rapor oluşturulamadı"
            }
    
    async def _prepare_report_data(self, score_data: Dict[str, Any], algorithm_run_id: Optional[int]) -> Dict[str, Any]:
        """Rapor verilerini hazırla"""
        return {
            "metadata": {
                "generated_at": datetime.now().isoformat(),
                "algorithm_run_id": algorithm_run_id,
                "report_title": "Proje Atama Optimizasyon Raporu",
                "report_subtitle": "Bitirme ve Ara Projeleri için Sınıf ve Danışman Atama Sistemi"
            },
            "summary": {
                "total_projects": score_data["performance_metrics"]["total_projects"],
                "total_instructors": score_data["performance_metrics"]["total_instructors"],
                "total_classrooms": score_data["performance_metrics"]["total_classrooms"],
                "total_timeslots": score_data["performance_metrics"]["total_timeslots"],
                "overall_score": score_data["weighted_total_score"]["score"],
                "optimization_quality": score_data["performance_metrics"]["optimization_quality"]
            },
            "objective_scores": score_data["objective_scores"],
            "detailed_analysis": score_data["detailed_analysis"],
            "recommendations": score_data["recommendations"],
            "schedules": await self._get_schedule_data(),
            "instructor_assignments": await self._get_instructor_assignments(),
            "classroom_utilization": await self._get_classroom_utilization(),
            "project_assignments": await self._get_project_assignments()
        }
    
    async def _generate_pdf_report(self, report_data: Dict[str, Any], algorithm_run_id: Optional[int]) -> Dict[str, Any]:
        """PDF raporu oluştur"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"optimization_report_{timestamp}.pdf"
            file_path = self.output_dir / filename
            
            # PDF dokümanı oluştur
            doc = SimpleDocTemplate(str(file_path), pagesize=A4)
            styles = getSampleStyleSheet()
            
            # Özel stil tanımları
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=TA_CENTER,
                textColor=colors.darkblue
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=12,
                textColor=colors.darkblue
            )
            
            # İçerik oluştur
            story = []
            
            # Başlık
            story.append(Paragraph(report_data["metadata"]["report_title"], title_style))
            story.append(Paragraph(report_data["metadata"]["report_subtitle"], styles['Normal']))
            story.append(Paragraph(f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Özet bölümü
            story.append(Paragraph("Genel Özet", heading_style))
            summary_table = self._create_summary_table(report_data["summary"])
            story.append(summary_table)
            story.append(Spacer(1, 20))
            
            # Amaç fonksiyonu skorları
            story.append(Paragraph("Amaç Fonksiyonu Skorları", heading_style))
            scores_table = self._create_scores_table(report_data["objective_scores"])
            story.append(scores_table)
            story.append(Spacer(1, 20))
            
            # Detaylı analiz
            story.append(Paragraph("Detaylı Analiz", heading_style))
            analysis_text = self._format_analysis_text(report_data["detailed_analysis"])
            story.append(Paragraph(analysis_text, styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Öneriler
            story.append(Paragraph("İyileştirme Önerileri", heading_style))
            recommendations_text = self._format_recommendations_text(report_data["recommendations"])
            story.append(Paragraph(recommendations_text, styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Hoca atamaları
            story.append(Paragraph("Hoca Atamaları", heading_style))
            instructor_table = self._create_instructor_assignments_table(report_data["instructor_assignments"])
            story.append(instructor_table)
            
            # PDF'i oluştur
            doc.build(story)
            
            file_size = file_path.stat().st_size
            
            return {
                "success": True,
                "file_path": str(file_path),
                "file_size": file_size,
                "message": "PDF raporu başarıyla oluşturuldu"
            }
            
        except Exception as e:
            logger.error(f"Error generating PDF report: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "message": "PDF raporu oluşturulamadı"
            }
    
    async def _generate_excel_report(self, report_data: Dict[str, Any], algorithm_run_id: Optional[int]) -> Dict[str, Any]:
        """Excel raporu oluştur"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"optimization_report_{timestamp}.xlsx"
            file_path = self.output_dir / filename
            
            # Excel workbook oluştur
            wb = Workbook()
            
            # Stil tanımları
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 1. Özet sayfası
            ws_summary = wb.active
            ws_summary.title = "Genel Özet"
            self._create_excel_summary_sheet(ws_summary, report_data["summary"], header_font, header_fill, border)
            
            # 2. Skorlar sayfası
            ws_scores = wb.create_sheet("Amaç Fonksiyonu Skorları")
            self._create_excel_scores_sheet(ws_scores, report_data["objective_scores"], header_font, header_fill, border)
            
            # 3. Hoca atamaları sayfası
            ws_instructors = wb.create_sheet("Hoca Atamaları")
            self._create_excel_instructor_sheet(ws_instructors, report_data["instructor_assignments"], header_font, header_fill, border)
            
            # 4. Sınıf kullanımı sayfası
            ws_classrooms = wb.create_sheet("Sınıf Kullanımı")
            self._create_excel_classroom_sheet(ws_classrooms, report_data["classroom_utilization"], header_font, header_fill, border)
            
            # 5. Proje atamaları sayfası
            ws_projects = wb.create_sheet("Proje Atamaları")
            self._create_excel_projects_sheet(ws_projects, report_data["project_assignments"], header_font, header_fill, border)
            
            # Excel dosyasını kaydet
            wb.save(str(file_path))
            
            file_size = file_path.stat().st_size
            
            return {
                "success": True,
                "file_path": str(file_path),
                "file_size": file_size,
                "message": "Excel raporu başarıyla oluşturuldu"
            }
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "message": "Excel raporu oluşturulamadı"
            }
    
    def _create_summary_table(self, summary: Dict[str, Any]) -> Table:
        """PDF için özet tablosu oluştur"""
        data = [
            ["Metrik", "Değer"],
            ["Toplam Proje Sayısı", str(summary["total_projects"])],
            ["Toplam Hoca Sayısı", str(summary["total_instructors"])],
            ["Toplam Sınıf Sayısı", str(summary["total_classrooms"])],
            ["Toplam Zaman Dilimi", str(summary["total_timeslots"])],
            ["Genel Skor", f"{summary['overall_score']:.2f}/100"],
            ["Optimizasyon Kalitesi", summary["optimization_quality"]]
        ]
        
        table = Table(data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        return table
    
    def _create_scores_table(self, objective_scores: Dict[str, Any]) -> Table:
        """PDF için skorlar tablosu oluştur"""
        data = [
            ["Amaç Fonksiyonu", "Skor", "Açıklama"]
        ]
        
        for score_name, score_data in objective_scores.items():
            data.append([
                score_name.replace("_", " ").title(),
                f"{score_data['score']:.2f}/100",
                score_data["description"][:50] + "..." if len(score_data["description"]) > 50 else score_data["description"]
            ])
        
        table = Table(data, colWidths=[2*inch, 1*inch, 3*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        return table
    
    def _create_instructor_assignments_table(self, instructor_assignments: List[Dict[str, Any]]) -> Table:
        """PDF için hoca atamaları tablosu oluştur"""
        data = [
            ["Hoca Adı", "Sınıf", "Zaman Dilimi", "Proje Sayısı", "Toplam Yük"]
        ]
        
        for assignment in instructor_assignments[:20]:  # İlk 20 kayıt
            data.append([
                assignment.get("instructor_name", "N/A"),
                assignment.get("classroom", "N/A"),
                assignment.get("timeslot", "N/A"),
                str(assignment.get("project_count", 0)),
                f"{assignment.get('total_load', 0):.1f}"
            ])
        
        if len(instructor_assignments) > 20:
            data.append(["...", "...", "...", "...", "..."])
        
        table = Table(data, colWidths=[1.5*inch, 1*inch, 1.5*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        return table
    
    def _format_analysis_text(self, detailed_analysis: Dict[str, Any]) -> str:
        """Analiz metnini formatla"""
        text = "Yük Dengesi Analizi:\n"
        text += f"• Mevcut standart sapma: {detailed_analysis.get('load_balance_analysis', {}).get('current_std_deviation', 'N/A')}\n"
        text += f"• Hedef standart sapma: {detailed_analysis.get('load_balance_analysis', {}).get('target_std_deviation', 'N/A')}\n\n"
        
        text += "Sınıf Kullanım Analizi:\n"
        text += f"• En çok kullanılan sınıf: {detailed_analysis.get('classroom_usage_analysis', {}).get('most_used_classroom', 'N/A')}\n"
        text += f"• En az kullanılan sınıf: {detailed_analysis.get('classroom_usage_analysis', {}).get('least_used_classroom', 'N/A')}\n"
        text += f"• Kullanım oranı: {detailed_analysis.get('classroom_usage_analysis', {}).get('utilization_rate', 'N/A')}\n\n"
        
        text += "Zaman Dilimi Analizi:\n"
        text += f"• Yoğun saatler: {', '.join(detailed_analysis.get('time_slot_analysis', {}).get('peak_hours', []))}\n"
        text += f"• Düşük kullanım saatleri: {', '.join(detailed_analysis.get('time_slot_analysis', {}).get('low_usage_hours', []))}\n"
        
        return text
    
    def _format_recommendations_text(self, recommendations: List[Dict[str, Any]]) -> str:
        """Öneriler metnini formatla"""
        text = ""
        for i, rec in enumerate(recommendations, 1):
            text += f"{i}. {rec.get('category', 'Genel')} ({rec.get('priority', 'Orta')} Öncelik):\n"
            text += f"   • {rec.get('description', 'Açıklama yok')}\n"
            text += f"   • Önerilen aksiyon: {rec.get('action', 'Aksiyon belirtilmemiş')}\n"
            text += f"   • Beklenen iyileştirme: {rec.get('expected_improvement', 'Belirsiz')}\n\n"
        
        return text
    
    def _create_excel_summary_sheet(self, ws, summary: Dict[str, Any], header_font, header_fill, border):
        """Excel özet sayfası oluştur"""
        ws['A1'] = "Genel Özet"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Veri ekle
        data = [
            ["Metrik", "Değer"],
            ["Toplam Proje Sayısı", summary["total_projects"]],
            ["Toplam Hoca Sayısı", summary["total_instructors"]],
            ["Toplam Sınıf Sayısı", summary["total_classrooms"]],
            ["Toplam Zaman Dilimi", summary["total_timeslots"]],
            ["Genel Skor", f"{summary['overall_score']:.2f}/100"],
            ["Optimizasyon Kalitesi", summary["optimization_quality"]]
        ]
        
        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if row_num == 2:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
    
    def _create_excel_scores_sheet(self, ws, objective_scores: Dict[str, Any], header_font, header_fill, border):
        """Excel skorlar sayfası oluştur"""
        ws['A1'] = "Amaç Fonksiyonu Skorları"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Header
        headers = ["Amaç Fonksiyonu", "Skor", "Açıklama", "Birim"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Veri ekle
        for row_num, (score_name, score_data) in enumerate(objective_scores.items(), 3):
            ws.cell(row=row_num, column=1, value=score_name.replace("_", " ").title())
            ws.cell(row=row_num, column=2, value=f"{score_data['score']:.2f}")
            ws.cell(row=row_num, column=3, value=score_data["description"])
            ws.cell(row=row_num, column=4, value=score_data.get("unit", "percentage"))
            
            for col_num in range(1, 5):
                ws.cell(row=row_num, column=col_num).border = border
    
    def _create_excel_instructor_sheet(self, ws, instructor_assignments: List[Dict[str, Any]], header_font, header_fill, border):
        """Excel hoca atamaları sayfası oluştur"""
        ws['A1'] = "Hoca Atamaları"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Header
        headers = ["Hoca Adı", "Sınıf", "Zaman Dilimi", "Proje Sayısı", "Toplam Yük"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Veri ekle
        for row_num, assignment in enumerate(instructor_assignments, 3):
            ws.cell(row=row_num, column=1, value=assignment.get("instructor_name", "N/A"))
            ws.cell(row=row_num, column=2, value=assignment.get("classroom", "N/A"))
            ws.cell(row=row_num, column=3, value=assignment.get("timeslot", "N/A"))
            ws.cell(row=row_num, column=4, value=assignment.get("project_count", 0))
            ws.cell(row=row_num, column=5, value=assignment.get("total_load", 0))
            
            for col_num in range(1, 6):
                ws.cell(row=row_num, column=col_num).border = border
    
    def _create_excel_classroom_sheet(self, ws, classroom_utilization: List[Dict[str, Any]], header_font, header_fill, border):
        """Excel sınıf kullanımı sayfası oluştur"""
        ws['A1'] = "Sınıf Kullanımı"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Header
        headers = ["Sınıf Adı", "Kullanım Sayısı", "Kullanım Oranı", "Boş Kalan Saatler"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Veri ekle
        for row_num, classroom in enumerate(classroom_utilization, 3):
            ws.cell(row=row_num, column=1, value=classroom.get("name", "N/A"))
            ws.cell(row=row_num, column=2, value=classroom.get("usage_count", 0))
            ws.cell(row=row_num, column=3, value=f"{classroom.get('utilization_rate', 0):.2%}")
            ws.cell(row=row_num, column=4, value=classroom.get("free_hours", "N/A"))
            
            for col_num in range(1, 5):
                ws.cell(row=row_num, column=col_num).border = border
    
    def _create_excel_projects_sheet(self, ws, project_assignments: List[Dict[str, Any]], header_font, header_fill, border):
        """Excel proje atamaları sayfası oluştur"""
        ws['A1'] = "Proje Atamaları"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Header
        headers = ["Proje ID", "Proje Başlığı", "Tür", "Sınıf", "Zaman", "Danışman", "Yardımcı 1", "Yardımcı 2"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Veri ekle
        for row_num, project in enumerate(project_assignments, 3):
            ws.cell(row=row_num, column=1, value=project.get("id", "N/A"))
            ws.cell(row=row_num, column=2, value=project.get("title", "N/A"))
            ws.cell(row=row_num, column=3, value=project.get("type", "N/A"))
            ws.cell(row=row_num, column=4, value=project.get("classroom", "N/A"))
            ws.cell(row=row_num, column=5, value=project.get("timeslot", "N/A"))
            ws.cell(row=row_num, column=6, value=project.get("responsible", "N/A"))
            ws.cell(row=row_num, column=7, value=project.get("second_participant", "N/A"))
            ws.cell(row=row_num, column=8, value=project.get("third_participant", "N/A"))
            
            for col_num in range(1, 9):
                ws.cell(row=row_num, column=col_num).border = border
    
    async def _get_schedule_data(self) -> List[Dict[str, Any]]:
        """Zamanlama verilerini getir (placeholder)"""
        return [
            {
                "timeslot": "09:00-09:30",
                "classrooms": ["D106", "D107"],
                "projects": 2,
                "instructors": 4
            },
            {
                "timeslot": "09:30-10:00",
                "classrooms": ["D106", "D107", "D108"],
                "projects": 3,
                "instructors": 6
            }
        ]
    
    async def _get_instructor_assignments(self) -> List[Dict[str, Any]]:
        """Hoca atamalarını getir (placeholder)"""
        return [
            {
                "instructor_name": "Dr. Ahmet Yılmaz",
                "classroom": "D106",
                "timeslot": "09:00-11:00",
                "project_count": 3,
                "total_load": 4.5
            },
            {
                "instructor_name": "Dr. Ayşe Demir",
                "classroom": "D107",
                "timeslot": "09:30-12:00",
                "project_count": 2,
                "total_load": 3.0
            }
        ]
    
    async def _get_classroom_utilization(self) -> List[Dict[str, Any]]:
        """Sınıf kullanımını getir (placeholder)"""
        return [
            {
                "name": "D106",
                "usage_count": 15,
                "utilization_rate": 0.85,
                "free_hours": "12:00-13:00"
            },
            {
                "name": "D107",
                "usage_count": 12,
                "utilization_rate": 0.72,
                "free_hours": "16:00-16:30"
            }
        ]
    
    async def _get_project_assignments(self) -> List[Dict[str, Any]]:
        """Proje atamalarını getir (placeholder)"""
        return [
            {
                "id": 1,
                "title": "Yapay Zeka Destekli Optimizasyon Sistemi",
                "type": "bitirme",
                "classroom": "D106",
                "timeslot": "09:00-09:30",
                "responsible": "Dr. Ahmet Yılmaz",
                "second_participant": "Dr. Ayşe Demir",
                "third_participant": "Arş. Gör. Mehmet Kaya"
            },
            {
                "id": 2,
                "title": "Mobil Uygulama Geliştirme",
                "type": "ara",
                "classroom": "D107",
                "timeslot": "09:30-10:00",
                "responsible": "Dr. Ayşe Demir",
                "second_participant": "Arş. Gör. Fatma Öz",
                "third_participant": "Dr. Ali Veli"
            }
        ]
